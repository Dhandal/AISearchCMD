import sys
import os
from bs4 import BeautifulSoup
import re 
import fitz
from datetime import datetime
import openpyxl
from os import path 
import time 
from PIL import Image
import requests
import pytesseract
import io
from io import BytesIO
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.http import MediaIoBaseDownload

log_file = open("message.log","w")
sys.stderr = log_file

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
creds = None
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
if os.path.exists('token.pickle'):
    with open('token.pickle', 'rb') as token:
        creds = pickle.load(token)
# If there are no (valid) credentials available, let the user log in.
if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)
    # Save the credentials for the next run
    with open('token.pickle', 'wb') as token:
        pickle.dump(creds, token)

service = build('drive', 'v3', credentials=creds)

def url_to_id(url):
    x = url.split("/")
    return x[5]

todayDate = datetime.now().strftime("%m/%d/%Y") 


questionList = []

className = sys.argv[1]
assignmentName = sys.argv[2]
folderToSearch = sys.argv[3]


#creates excel sheet 
if path.exists(className +".xlsx"):
    wb = openpyxl.load_workbook(className + ".xlsx")
else:
    wb = openpyxl.Workbook()

if assignmentName not in wb.sheetnames:
    wb.create_sheet(assignmentName)

#gets rid of initial sheet
if "Sheet" in wb.sheetnames:
    wb.remove_sheet(wb["Sheet"])

sheet = wb[assignmentName]

#sets up sheet 
if sheet.cell(row =1, column=1).value is None:
        sheet.cell(row =1, column=1).value = "QUESTION #"
        sheet.cell(row =1, column=2).value = "QUESTION"
        sheet.cell(row =1, column=3).value = "LINK"
        sheet.cell(row =1, column=4).value = "DATE"
        sheet.cell(row =1, column=5).value = "FOUND BEFORE"
maxRow = sheet.max_row



for file in os.listdir(folderToSearch):
    if file.endswith('.html'):
        f = open(os.path.join(folderToSearch,file), 'r')
        contents = f.read()
        soup = BeautifulSoup(contents,'lxml')
        questions = soup.find_all('p')
        for question in questions: 
            questionText = question.find_all("p")   
            completeQuestion = ""
            for text in questionText:
                #get pdf if there is a pdf
                link = text.find('a')
                #print(link)
                if link is not None:
                    link = link.get('href')
                    if "drive.google" in link:
                        driveId = url_to_id(link)
                        file = service.files().get(fileId=driveId).execute()
                        extension = service.files().get(fileId=driveId, fields="fileExtension").execute()['fileExtension']
                        dest = 'temp.' + extension
                        request = service.files().get_media(fileId=driveId)
                        fh = io.FileIO(dest,'w')
                        downloader = MediaIoBaseDownload(fh, request)
                        done = False
                        while done is False:
                            done = downloader.next_chunk()
                            
                        if 'pdf' in dest:
                            doc = fitz.open(dest)  # open document
                            page = doc[0]
                            for page in doc: 
                                text = page.getText()
                                completeQuestion = completeQuestion + text
                        elif 'png' in dest or 'jpg' in dest:
                            img = Image.open(dest)
                            text = pytesseract.image_to_string(img)
                            completeQuestion = completeQuestion + text 
                        elif 'txt' in dest:
                            readFile = open(dest,'r')
                            completeQuestion = completeQuestion + readFile.read()
                    else:
                        pdfLink = link
                        #downloading pdf and getting data
                        r = requests.get(pdfLink,stream = True)
                        with open('metadata.pdf', 'wb') as fd:
                            for chunk in r.iter_content(2000):
                                fd.write(chunk)
                        doc = fitz.open('metadata.pdf')  # open document
                        page = doc[0]
                        for page in doc: 
                            text = page.getText()
                            completeQuestion = completeQuestion + text
                        fd.close()
                #get img if there is an image
                elif text.find('img') is not None:
                    imgLink = text.find('img').get('src')
                    print(imgLink)
                    if not imgLink.startswith('https'):
                        if imgLink.startswith('/'):
                            imgLink = 'https://canvas.ucsd.edu' + str(imgLink)
                        else:
                            imgLink = 'https://canvas.ucsd.edu/' + str(imgLink)
                    
                    r = requests.get(imgLink)
                    img = Image.open(BytesIO(r.content))
                    text = pytesseract.image_to_string(img)
                    completeQuestion = completeQuestion + text     
                
                else:
                    completeQuestion = completeQuestion + text.text
                if os.path.exists('metadata.pdf'):
                    os.remove('metadata.pdf')
            
            questionList.append(completeQuestion)
    elif file.endswith('.txt'):
        f = open(os.path.join(folderToSearch,file), 'r') 
        line = f.readline()
        while line:
            if re.match(r"\d\)+|\d_|+[\u2022,\u2023,\u25E6,\u2043,\u2219]+",line):
                questionList.append(line)
            else: 
                questionList[len(questionList-1)] = questionList[len(questionList-1)] + line
            line = f.readline()
    elif file.endswith('.pdf'):
        doc = fitz.open(os.path.join(folderToSearch,file))
        page = doc[0]
        for page in doc:
            text = page.getText()
            temp = open('temp.txt','w')
            temp.write(text)
        line = temp.readline()
        while line:
            if re.match(r"\d\)+|\d_|+[\u2022,\u2023,\u25E6,\u2043,\u2219]+",line):
                questionList.append(line)
            else: 
                questionList[len(questionList-1)] = questionList[len(questionList-1)] + line
            line = f.readline()
        if os.path.exists('temp.txt'):
            os.remove('temp.txt')


print(len(questionList))
for question in questionList:
    print(question)
    
"""
try: 
    from googlesearch import search 
except ImportError:  
    print("No module named 'google' found") 
#keeps track of current question number 
qNum = 0
for question in questionList:
    qNum = qNum + 1
    query = question
    #searches google for any chegg links within the first 10 results    
    for j in search(query, tld="co.in", num=10, stop=10, pause=1): 
        if "chegg" in j or "coursehero" in j:
            
            URL = j
            newRow = [qNum, query,URL,todayDate]
            #used to check for duplicates
            found = False 
            #used to check if the question has been entered before
            foundBefore = False
            #checks to see if row to be added already exists 
            for row in sheet.iter_rows(max_col = 4,max_row = maxRow):  
                data = []
                for cell in row:
                    data.append(cell.value)
                if data == newRow:
                    found = True
                    break  
            if not found:
                #checks to see if question has been added before
                for row in sheet.iter_rows(max_row = maxRow, min_col=2,max_col=2):
                    for cell in row:
                        if cell.value == question:
                            foundBefore = True
                            break 
                newRow.append(str(foundBefore))
                sheet.append(tuple(newRow)) 
                currentCell ='C' + str(sheet.max_row) 
                sheet[currentCell].style = "Hyperlink"           

    print("Questions " + "(" + str(qNum) + "/" + str(len(questionList)) + ") " + "done")
print("done!")
"""
dims = {}
for row in sheet.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
for col, value in dims.items():
    sheet.column_dimensions[col].width = value
time.sleep(2)
wb.save(className + ".xlsx")

